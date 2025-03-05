import os
import pyodbc
import openpyxl
from resourcesMail2 import create_connectionMail, read_properties
from resourcesDB import create_connection
import logging
from Job import Job
from contextlib import closing

class MailQueryAllegatoExcelOpenpyxl(Job):
    def run(self):
        try:
            self.logger.info(f"Esecuzione del job: {self.NomeJob}")

            if not os.path.exists(self.excel_path):
                os.makedirs(self.excel_path)
                self.logger.info(f"Creata directory: {self.excel_path}")

            query = self.read_query(self.query_path)
            self.logger.info(f"Query da eseguire: {query}")
            
            is_html = self.is_html if self.is_html is not None else False
            
            with closing(create_connection(self.config_path_db)) as connection:
                self.logger.info("Connessione al database stabilita con successo.")                
                try:
                    cursor = connection.cursor()
                    cursor.execute(query)
                    
                    if cursor.description is None:
                        self.logger.error("La query non ha restituito informazioni sulle colonne")
                        return
                    
                    columns = [column[0] for column in cursor.description]
                    results = cursor.fetchall()
                    self.logger.info(f"Query eseguita con successo. Righe recuperate: {len(results)}")
                    self.logger.debug(f"Prime righe: {results[:5] if results else 'Nessuna riga'}")
                except pyodbc.Error as e:
                    self.logger.error(f"Errore del database: {e}")
                    print(f"Errore del database: {e}")
                    return

            if not results:
                self.logger.info("La query non ha restituito righe, invio email senza allegato.")
                self.body += "\n\nNessun record trovato."
                properties_mail = read_properties(self.config_path_mail)
                emailReplay = properties_mail['emailReplay']
                emailFrom = properties_mail['emailFrom']
                smtpServer = properties_mail['smtpServer']                
                create_connectionMail(smtpServer, emailFrom, emailReplay, self.subject, self.body, self.to_email, self.cc_email, is_html=is_html)
                self.logger.info("Email inviata con successo.")
                print("Email inviata con successo.")
                return

            excel_path = os.path.join(self.excel_path, self.file_excel)

            # Creazione del file Excel usando openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active

            # Scrittura delle intestazioni
            for col_num, column_title in enumerate(columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)

            # Scrittura dei dati
            for row_num, row in enumerate(results, 2):
                for col_num, cell_value in enumerate(row, 1):
                    if cell_value is not None:
                        if isinstance(cell_value, (int, float, str, bool)):
                            ws.cell(row=row_num, column=col_num, value=cell_value)
                        else:
                            ws.cell(row=row_num, column=col_num, value=str(cell_value))
                    else:
                        ws.cell(row=row_num, column=col_num, value="")

            wb.save(excel_path)
            self.logger.info("Risultati della query salvati su Excel.")

            properties_mail = read_properties(self.config_path_mail)
            emailReplay = properties_mail['emailReplay']
            emailFrom = properties_mail['emailFrom']            
            smtpServer = properties_mail['smtpServer']
            create_connectionMail(smtpServer, emailFrom, emailReplay, self.subject, self.body, self.to_email, self.cc_email, [excel_path], is_html=is_html)
            self.logger.info("Email con allegato inviata con successo.")
            print("Email inviata con successo.")

        except Exception as e:
            self.logger.error(f"Errore durante l'esecuzione del job: {e}")
            print(f"Errore durante l'esecuzione del job: {e}")

    def read_query(self, query_path):
        try:
            with open(query_path, 'r') as file:
                query = file.read()
            return query
        except IOError as e:
            self.logger.error(f"Errore nella lettura del file della query: {e}")
            raise