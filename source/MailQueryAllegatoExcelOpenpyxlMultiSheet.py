import os
import pyodbc
import openpyxl
import re
from resourcesMail2 import create_connectionMail, read_properties
from resourcesDB import create_connection
import logging
from Job import Job
from contextlib import closing

class MailQueryAllegatoExcelOpenpyxlMultiSheet(Job):
    def run(self):
        try:
            self.logger.info(f"Esecuzione del job: {self.NomeJob}")

            if not os.path.exists(self.excel_path):
                os.makedirs(self.excel_path)
                self.logger.info(f"Creata directory: {self.excel_path}")

            queries_info = self.read_queries_from_csv(self.query_path)
            if not queries_info:
                self.logger.error(f"Impossibile leggere le query dal file {os.path.basename(self.query_path)}")
                return

            config_date = self.read_date_from_config(self.config_path)
            
            if not config_date:			   
                self.logger.info(f"Data di riferimento:{self.get_eomonth_date()}")
                data_mail=self.get_eomonth_date()                
            else: 
                data_mail=config_date               
                self.logger.info(f"Data di riferimento: {config_date}")

            is_html = self.is_html if self.is_html is not None else False
            
            excel_path = os.path.join(self.excel_path, self.file_excel)
            wb = openpyxl.Workbook()
            
            with closing(create_connection(self.config_path_db)) as connection:
                self.logger.info("Connessione al database stabilita con successo.")
                
                for sheet_name, query in queries_info:
                    prepared_query = self.prepare_query(query, config_date)
                    self.logger.info(f"Esecuzione query per sheet: {sheet_name}")
                    
                    try:
                        cursor = connection.cursor()
                        cursor.execute(prepared_query)
                        
                        if cursor.description is None:
                            self.logger.error(f"La query per '{sheet_name}' non ha restituito informazioni sulle colonne")
                            continue
                        
                        columns = [column[0] for column in cursor.description]
                        results = cursor.fetchall()
                        
                        self.create_sheet(wb, sheet_name, columns, results)
                        if results:
                            self.logger.info(f"Query eseguita con successo per '{sheet_name}'. Righe recuperate: {len(results)}")
                        else:
                            self.logger.info(f"La query per '{sheet_name}' non ha restituito righe")
                    
                    except pyodbc.Error as e:
                        self.logger.error(f"Errore del database per '{sheet_name}': {e}")
                        print(f"Errore del database per '{sheet_name}': {e}")

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            wb.save(excel_path)
            self.logger.info("Risultati delle query salvati su Excel.")
            self.send_email_with_attachment(excel_path, is_html,data_mail)

        except Exception as e:
            self.logger.error(f"Errore durante l'esecuzione del job: {e}")
            print(f"Errore durante l'esecuzione del job: {e}")

    def read_queries_from_csv(self, file_path):
        try:
            queries = []
            with open(file_path, 'r', newline='', encoding='utf-8') as file:
                content = file.read()
                
                # Ignorare la prima riga (intestazioni)
                content = content.split('\n', 1)[1]
                
                # Usare espressioni regolari per trovare le query tra virgolette
                pattern = r'([^,\n]+),\s*"((?:[^"\\]|\\.)*)"'
                matches = re.findall(pattern, content, re.DOTALL)
                
                for sheet_name, query in matches:
                    sheet_name = sheet_name.strip()
                    query = query.strip()
                    # Rimuovere le virgolette di escape all'interno della query
                    query = query.replace('\\"', '"')
                    queries.append((sheet_name, query))
            
            self.logger.info(f"Lette {len(queries)} query dal file CSV")
            for idx, (sheet_name, query) in enumerate(queries, 1):
                self.logger.info(f"Query {idx}: Sheet '{sheet_name}', primi 100 caratteri: {query[:100]}...")
                self.logger.debug(f"Query {idx} completa:\n{query}")
            
            return queries
        except IOError as e:
            self.logger.error(f"Errore nella lettura del file {os.path.basename(file_path)}: {e}")
            return None
        except Exception as e:
            self.logger.error(f"Errore durante l'elaborazione del file CSV: {e}")
            return None

    def read_date_from_config(self, config_path):
        try:
            with open(config_path, 'r') as file:
                for line in file:
                    if line.startswith('reference_date'):
                        return line.split('=')[1].strip()
            self.logger.error(f"reference_date not found in {config_path}")
            return None
        except Exception as e:
            self.logger.error(f"Errore nella lettura della data dal file {os.path.basename(config_path)}: {e}")
            return None
    def get_eomonth_date(self):
        
        try:
            with closing(create_connection(self.config_path_db)) as connection:
                cursor = connection.cursor()
                cursor.execute("SELECT CONVERT(VARCHAR(10), EOMONTH(GETDATE(), -1), 120)")
                eomonth_date = cursor.fetchone()[0]
                #self.logger.info(f"Utilizzo EOMONTH(GETDATE(), -1) come data di riferimento: {eomonth_date}")
                return eomonth_date
        except Exception as e:
            self.logger.error(f"Errore durante il recupero di EOMONTH dal database: {e}")
            return None
			
    def prepare_query(self, query, config_date):
        if config_date:
            return query.replace("@FirstDayOfPreviousMonth", f"'{config_date}'")
        else:
            # Se per qualche motivo non abbiamo una data, utilizziamo EOMONTH direttamente nella query
            return query.replace("@FirstDayOfPreviousMonth", "EOMONTH(GETDATE(), -1)")

    def create_sheet(self, workbook, sheet_title, columns, results):
        ws = workbook.create_sheet(title=sheet_title)
        
        # Scrittura delle intestazioni
        for col_num, column_title in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)

        if results is None or len(results) == 0:
            # Aggiungi un messaggio quando non ci sono risultati
            ws.cell(row=2, column=1, value="La query non ha restituito righe")
        else:
            # Scrittura dei dati
            for row_num, row in enumerate(results, 2):
                for col_num, cell_value in enumerate(row, 1):
                    if cell_value is not None:
                        if isinstance(cell_value, (int, float, str, bool)):
                            ws.cell(row=row_num, column=col_num, value=cell_value)
                        else:
                            ws.cell(row=row_num, column=col_num, value=str(cell_value))
                    else:
                        ws.cell(row=row_num, column=col_num, value="")

    def send_email_with_attachment(self, excel_path, is_html,data):
        subject_with_date = f"{self.subject} - fatturazione competenza: {data}" 
        properties_mail = read_properties(self.config_path_mail)
        emailReplay = properties_mail['emailReplay']
        emailFrom = properties_mail['emailFrom']            
        smtpServer = properties_mail['smtpServer']
        create_connectionMail(smtpServer, emailFrom, emailReplay, subject_with_date, self.body, self.to_email, self.cc_email, [excel_path], is_html=is_html)
        self.logger.info("Email con allegato inviata con successo.")
        print("Email inviata con successo con allegato.")
